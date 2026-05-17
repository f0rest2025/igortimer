import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

def export_to_excel(db, file_path):
    """
    Exports all time segments to an Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time Tracker Report"

    # Headers
    headers = [
        "Дата", "Клиент", "Задача", "Начало", "Конец", 
        "Длительность (мин)", "Статус", "Заметки"
    ]
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Fetch data
    segments = db.get_segments()
    
    for seg in segments:
        start_dt = datetime.fromisoformat(seg['start_at'])
        end_dt = datetime.fromisoformat(seg['end_at']) if seg['end_at'] else datetime.now()
        
        duration = (end_dt - start_dt).total_seconds() / 60
        
        row_data = [
            start_dt.strftime("%Y-%m-%d"),
            seg['client_name'] or "N/A",
            seg['task'],
            start_dt.strftime("%H:%M:%S"),
            end_dt.strftime("%H:%M:%S") if seg['end_at'] else "Активен",
            round(duration, 2),
            seg['status'],
            seg['note'] or ""
        ]
        ws.append(row_data)

    # Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    wb.save(file_path)
    return True
